import datetime
import re
from io import BytesIO, StringIO

import pandas as pd
import pytz
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import numbers


SHEET_NAME = "Data"
USER_ID_SHEET = "UserID"
CORE_COLS = ["Sample_ID", "I7_Index_ID", "index", "I5_Index_ID", "index2", "Amplicon", "gRNA"]
TEXT_BLANK_COLS = ["ELN_ID", "Isoform_Sample_ID", "PAM", "Base_Editing_Type"]
FINAL_COLS = [
    "Sample_ID", "Sample_Name", "I7_Index_ID", "index", "I5_Index_ID", "index2",
    "Sample_Project", "Description", "ELN_ID", "Isoform_Sample_ID", "PAM",
    "gRNA", "Amplicon", "Exon",
    "Expected_HDR_Amplicon", "Quantification_Window_Coordinates",
    "Quantification_Window_Center", "Plot_Window_Size", "ngRNA", "Base_Editing_Type"
]
NGRNA_RAW_HEADERS = ["ngRNA\n(nicking RNA)", "ngRNA (nicking RNA)"]
NGRNA_FINAL_HEADER = "ngRNA"
BASE_EDITING_COL = "Base_Editing_Type"
VALID_BASE_EDITING_TYPES = {"ABE", "CBE", "BOTH"}
BASE_EDITING_CONFLICT_COLS = [
    "Expected_HDR_Amplicon",
    "Quantification_Window_Coordinates",
    "Quantification_Window_Center",
    "Plot_Window_Size",
    NGRNA_FINAL_HEADER,
]
DNA_ONLY_RE = re.compile(r"^[ATCG]*$")
DNA_OR_U_RE = re.compile(r"^[ATCGU]*$")
QC_ANCHOR_COLS = ["Sample_ID", "gRNA", "Amplicon", "Exon", "Expected_HDR_Amplicon", BASE_EDITING_COL]
QC_OPTIONAL_COLS = [
    "Exon", "Expected_HDR_Amplicon", "Quantification_Window_Coordinates",
    "Quantification_Window_Center", "Plot_Window_Size", BASE_EDITING_COL, NGRNA_FINAL_HEADER,
]


st.set_page_config(page_title="CRISPResso QC / MiSeq Merge Tool", layout="wide")
st.title("🧪 CRISPResso QC / MiSeq Merge Tool")

pst = pytz.timezone("America/Los_Angeles")
today = datetime.datetime.now(pst).date()
default_prefix = f"run{today.year % 100:02d}{today.month:02d}{today.day:02d}mi"

mode = st.radio(
    "Mode",
    options=["QC Screening", "File Merge"],
    index=0,
    horizontal=True,
    help="QC Screening is read-only and reports issues. File Merge generates CRISPResso and MiSeq outputs.",
)

prefix = st.text_input("Filename prefix (used for File Merge outputs)", value=default_prefix)
out_excel = f"{prefix}_sample_info.xlsx"
out_csv = f"{prefix}_miseq.csv"

state = st.session_state
if "upload_key" not in state:
    state.upload_key = 0
for key in ("qc_results", "merge_results"):
    if key not in state:
        state[key] = None

col1, col2, _ = st.columns([1, 1, 4])
with col1:
    run_clicked = st.button("▶️ Run QC" if mode == "QC Screening" else "▶️ Merge")
with col2:
    if st.button("🗑️ Clear uploads"):
        state.upload_key += 1
        state.qc_results = None
        state.merge_results = None

st.markdown("---")

uploaded_files = st.file_uploader(
    "Upload one or more .xlsx files",
    type=["xlsx"],
    accept_multiple_files=True,
    key=f"uploads_{state.upload_key}",
)


def normalize_header_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for raw_name in NGRNA_RAW_HEADERS:
        if raw_name in df.columns and NGRNA_FINAL_HEADER not in df.columns:
            df = df.rename(columns={raw_name: NGRNA_FINAL_HEADER})
    return df


def clean_cell_minimal(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def blankify_minimal(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df = df.where(pd.notna(df), "")
    return df.applymap(clean_cell_minimal)


def clean_cell_merge(value) -> str:
    if value is None or pd.isna(value):
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none", "null", "0"}:
        return ""
    return s


def blankify_merge(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df = df.where(pd.notna(df), "")
    return df.applymap(clean_cell_merge)


def normalize_dna(value) -> str:
    return clean_cell_minimal(value).upper()


def normalize_grna_for_merge(value) -> str:
    s = clean_cell_merge(value)
    if not s:
        return ""
    return s.upper().replace("U", "T")


def normalize_base_editing_type(value) -> str:
    s = clean_cell_minimal(value)
    return s.upper() if s else ""


def reverse_complement(seq: str) -> str:
    s = clean_cell_minimal(seq).upper().replace("U", "T")
    trans = str.maketrans("ATCG", "TAGC")
    return s.translate(trans)[::-1]


def ensure_columns_and_order(df: pd.DataFrame, columns) -> pd.DataFrame:
    df = df.copy()
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    df = df[columns]
    df = blankify_merge(df)
    for col in TEXT_BLANK_COLS:
        if col in df.columns:
            df[col] = df[col].apply(clean_cell_merge)
    return df


def write_excel_with_blanks(df: pd.DataFrame, excel_buf: BytesIO, sheet_name: str, user_ids=None):
    df_to_write = df.replace("", None)
    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        df_to_write.to_excel(writer, index=False, sheet_name=sheet_name)
        if user_ids:
            pd.DataFrame({"UserID": [", ".join(user_ids)]}).to_excel(writer, index=False, sheet_name=USER_ID_SHEET)

    excel_buf.seek(0)
    wb = load_workbook(excel_buf)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.number_format = numbers.FORMAT_TEXT
            if cell.value == 0 or cell.value is None:
                cell.value = None

    excel_buf.seek(0)
    excel_buf.truncate()
    wb.save(excel_buf)
    excel_buf.seek(0)


def read_user_ids(excel_file) -> list:
    user_ids = set()
    try:
        df_user_id = pd.read_excel(
            excel_file,
            sheet_name=USER_ID_SHEET,
            engine="openpyxl",
            dtype=str,
            keep_default_na=False,
            na_filter=False,
        )
    except ValueError:
        return []

    cols_lower = [str(c).strip().replace("_", " ").lower() for c in df_user_id.columns]
    target_idx = next((i for i, c in enumerate(cols_lower) if c in {"userid", "user id"}), -1)
    if target_idx == -1:
        return []

    target_col = df_user_id.columns[target_idx]
    for value in df_user_id[target_col].astype(str).str.strip():
        if value:
            user_ids.add(value)
    return sorted(user_ids)


def read_data_sheet(excel_file):
    return pd.read_excel(
        excel_file,
        sheet_name=SHEET_NAME,
        engine="openpyxl",
        dtype=str,
        keep_default_na=False,
        na_filter=False,
    )


def display_file_label(excel_file, idx: int) -> str:
    return f"{excel_file.name} [upload {idx + 1}]"


def is_effectively_blank_for_activity(value) -> bool:
    s = clean_cell_minimal(value)
    return s in {"", "0"}


def row_is_effectively_blank(row: pd.Series, columns) -> bool:
    for col in columns:
        if col in row.index and not is_effectively_blank_for_activity(row.get(col, "")):
            return False
    return True


def add_issue(issue_rows: list, file_name: str, row_num: str, sample_id: str, severity: str, category: str, message: str):
    issue_rows.append({
        "File": file_name,
        "Row": row_num,
        "Sample_ID": sample_id,
        "Severity": severity,
        "Category": category,
        "Message": message,
    })


def analyze_duplicates(active_df: pd.DataFrame, file_name: str, issue_rows: list):
    details = []
    if active_df.empty or "index" not in active_df.columns or "index2" not in active_df.columns:
        return details, set()

    working = active_df.copy()
    working["_index_key"] = working["index"].apply(clean_cell_minimal)
    working["_index2_key"] = working["index2"].apply(clean_cell_minimal)
    valid_mask = (
        working["_index_key"].map(lambda v: not is_effectively_blank_for_activity(v))
        & working["_index2_key"].map(lambda v: not is_effectively_blank_for_activity(v))
    )
    dup_rows = working.loc[valid_mask].loc[
        working.loc[valid_mask].duplicated(subset=["_index_key", "_index2_key"], keep=False)
    ]

    combos = set(tuple(x) for x in working.loc[valid_mask, ["_index_key", "_index2_key"]].drop_duplicates().itertuples(index=False, name=None))

    if dup_rows.empty:
        return details, combos

    for (idx1, idx2), grp in dup_rows.groupby(["_index_key", "_index2_key"], sort=False):
        rows = grp["_row_num"].astype(str).tolist()
        samples = [s for s in grp["Sample_ID"].astype(str).tolist() if s]
        message = f"Duplicate index/index2 combination {idx1}/{idx2} found in rows: {', '.join(rows)}"
        for _, rec in grp.iterrows():
            add_issue(
                issue_rows,
                file_name=file_name,
                row_num=str(rec["_row_num"]),
                sample_id=clean_cell_minimal(rec.get("Sample_ID", "")),
                severity="Error",
                category="Index Duplication",
                message=message,
            )
        details.append({
            "File": file_name,
            "index": idx1,
            "index2": idx2,
            "Rows": ", ".join(rows),
            "Sample_IDs": ", ".join(samples),
        })
    return details, combos


def cross_file_duplicates(file_combo_map: dict):
    combo_files_map = {}
    for file_label, combos in file_combo_map.items():
        for combo in combos:
            combo_files_map.setdefault(combo, []).append(file_label)
    return {combo: files for combo, files in combo_files_map.items() if len(files) > 1}


def qc_screen_file(excel_file, upload_idx: int):
    file_name = display_file_label(excel_file, upload_idx)
    user_ids = read_user_ids(excel_file)
    try:
        df_raw = read_data_sheet(excel_file)
    except ValueError:
        return {
            "file_name": file_name,
            "sheet_found": False,
            "user_ids": user_ids,
            "issues": [{
                "File": file_name,
                "Row": "",
                "Sample_ID": "",
                "Severity": "Error",
                "Category": "Input / Structural QC",
                "Message": f'Missing required sheet "{SHEET_NAME}".',
            }],
            "amplicon_summary": [],
            "duplicate_details": [],
            "row_count_checked": 0,
            "combos": set(),
            "missing_required_columns": CORE_COLS.copy(),
        }

    df = normalize_header_columns(blankify_minimal(df_raw))
    if BASE_EDITING_COL not in df.columns:
        df[BASE_EDITING_COL] = ""
    if "Exon" not in df.columns:
        df["Exon"] = ""
    if "Expected_HDR_Amplicon" not in df.columns:
        df["Expected_HDR_Amplicon"] = ""
    if NGRNA_FINAL_HEADER not in df.columns:
        df[NGRNA_FINAL_HEADER] = ""
    df["_row_num"] = df.index + 2

    issues = []
    missing_required_columns = [col for col in CORE_COLS if col not in df.columns]
    for col in missing_required_columns:
        add_issue(
            issues, file_name, "", "", "Error", "Input / Structural QC",
            f'Missing required column "{col}".',
        )

    rows_checked = 0
    unique_amplicons = []
    amp_seen = {}

    relevant_cols = list(dict.fromkeys(CORE_COLS + QC_OPTIONAL_COLS))
    available_relevant_cols = [c for c in relevant_cols if c in df.columns]

    active_df = df.loc[~df.apply(lambda row: row_is_effectively_blank(row, available_relevant_cols), axis=1)].copy()

    for _, row in active_df.iterrows():
        rows_checked += 1
        row_num = str(int(row["_row_num"]))
        sample_id = clean_cell_minimal(row.get("Sample_ID", ""))

        missing_fields = []
        for col in CORE_COLS:
            if col not in df.columns:
                continue
            if clean_cell_minimal(row.get(col, "")) == "":
                missing_fields.append(col)
        if missing_fields:
            add_issue(
                issues, file_name, row_num, sample_id, "Error", "Core Row QC",
                f"Missing required value(s): {', '.join(missing_fields)}",
            )

        amp_raw = clean_cell_minimal(row.get("Amplicon", ""))
        amp_norm = normalize_dna(amp_raw)
        if amp_norm:
            if amp_norm not in amp_seen:
                amp_seen[amp_norm] = f"Amplicon{len(amp_seen)+1:02d}"
                unique_amplicons.append({
                    "Amplicon Label": amp_seen[amp_norm],
                    "Length": len(amp_norm),
                    "Occurrence Count": 1,
                    "Sequence": amp_norm,
                })
            else:
                for item in unique_amplicons:
                    if item["Sequence"] == amp_norm:
                        item["Occurrence Count"] += 1
                        break

            if "U" in amp_norm:
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Amplicon QC",
                    "Amplicon contains U. Amplicon must be a DNA sequence with A/T/C/G only.",
                )
            elif not DNA_ONLY_RE.fullmatch(amp_norm):
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Amplicon QC",
                    "Amplicon contains characters other than A/T/C/G.",
                )

        grna_raw = clean_cell_minimal(row.get("gRNA", ""))
        grna_norm = normalize_dna(grna_raw)
        if grna_norm:
            if "U" in grna_norm:
                add_issue(
                    issues, file_name, row_num, sample_id, "Info", "gRNA QC",
                    "gRNA contains U. This will be converted to T in File Merge mode.",
                )
            elif not DNA_ONLY_RE.fullmatch(grna_norm) and not DNA_OR_U_RE.fullmatch(grna_norm):
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "gRNA QC",
                    "gRNA contains characters other than A/T/C/G/U.",
                )

            if amp_norm and DNA_ONLY_RE.fullmatch(amp_norm) and DNA_OR_U_RE.fullmatch(grna_norm):
                grna_for_match = grna_norm.replace("U", "T")
                grna_rc = reverse_complement(grna_for_match)
                if grna_for_match not in amp_norm and grna_rc not in amp_norm:
                    add_issue(
                        issues, file_name, row_num, sample_id, "Error", "Amplicon QC",
                        "gRNA was not found in Amplicon, and reverse complement(gRNA) was also not found.",
                    )

        exon_raw = clean_cell_minimal(row.get("Exon", ""))
        exon_norm = normalize_dna(exon_raw)
        if exon_norm:
            if "U" in exon_norm:
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Exon QC",
                    "Exon contains U. Exon must be a DNA sequence with A/T/C/G only.",
                )
            elif not DNA_ONLY_RE.fullmatch(exon_norm):
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Exon QC",
                    "Exon contains characters other than A/T/C/G.",
                )
            elif amp_norm and DNA_ONLY_RE.fullmatch(amp_norm) and exon_norm not in amp_norm:
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Exon QC",
                    "Exon was not found in Amplicon in the same orientation.",
                )

        hdr_amp_raw = clean_cell_minimal(row.get("Expected_HDR_Amplicon", ""))
        hdr_amp_norm = normalize_dna(hdr_amp_raw)
        if hdr_amp_norm:
            if "U" in hdr_amp_norm:
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Logical Consistency QC",
                    "Expected_HDR_Amplicon contains U. Expected_HDR_Amplicon must be a DNA sequence with A/T/C/G only.",
                )
            elif not DNA_ONLY_RE.fullmatch(hdr_amp_norm):
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Logical Consistency QC",
                    "Expected_HDR_Amplicon contains characters other than A/T/C/G.",
                )
            if clean_cell_minimal(row.get("Quantification_Window_Coordinates", "")) == "":
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Logical Consistency QC",
                    "Expected_HDR_Amplicon is filled but Quantification_Window_Coordinates is blank.",
                )

        base_edit_val = normalize_base_editing_type(row.get(BASE_EDITING_COL, ""))
        if base_edit_val:
            if base_edit_val not in VALID_BASE_EDITING_TYPES:
                add_issue(
                    issues, file_name, row_num, sample_id, "Error", "Base Editing QC",
                    f'Invalid Base_Editing_Type "{base_edit_val}". Allowed values are ABE, CBE, BOTH.',
                )
            else:
                has_conflict = any(clean_cell_minimal(row.get(col, "")) != "" for col in BASE_EDITING_CONFLICT_COLS if col in df.columns)
                if has_conflict:
                    add_issue(
                        issues, file_name, row_num, sample_id, "Error", "Base Editing QC",
                        "Base-editing row also contains HDR/window/ngRNA fields.",
                    )

    dup_details, combos = analyze_duplicates(active_df, file_name, issues)

    return {
        "file_name": file_name,
        "sheet_found": True,
        "user_ids": user_ids,
        "issues": issues,
        "amplicon_summary": unique_amplicons,
        "duplicate_details": dup_details,
        "row_count_checked": rows_checked,
        "combos": combos,
        "missing_required_columns": missing_required_columns,
    }


def run_qc(uploaded_files):
    all_issues = []
    amplicon_rows = []
    duplicate_rows = []
    user_ids = set()
    file_summaries = []
    file_combo_map = {}
    rows_checked_total = 0

    for upload_idx, excel_file in enumerate(uploaded_files):
        result = qc_screen_file(excel_file, upload_idx)
        file_name = result["file_name"]
        rows_checked_total += result["row_count_checked"]
        user_ids.update(result["user_ids"])
        file_combo_map[file_name] = result["combos"]
        all_issues.extend(result["issues"])
        for row in result["amplicon_summary"]:
            amplicon_rows.append({"File": file_name, **row})
        duplicate_rows.extend(result["duplicate_details"])

        issue_df = pd.DataFrame(result["issues"])
        errors = int((issue_df["Severity"] == "Error").sum()) if not issue_df.empty else 0
        infos = int((issue_df["Severity"] == "Info").sum()) if not issue_df.empty else 0
        warnings = int((issue_df["Severity"] == "Warning").sum()) if not issue_df.empty else 0
        file_summaries.append({
            "File": file_name,
            "Rows Checked": result["row_count_checked"],
            "Errors": errors,
            "Warnings": warnings,
            "Info": infos,
            "Status": "Needs Correction" if errors > 0 else "Pass",
        })

    cross_dup = cross_file_duplicates(file_combo_map)
    for (idx1, idx2), files in cross_dup.items():
        message = f"Cross-file duplicate index/index2 combination {idx1}/{idx2} found in files: {', '.join(files)}"
        all_issues.append({
            "File": ", ".join(files),
            "Row": "",
            "Sample_ID": "",
            "Severity": "Error",
            "Category": "Index Duplication",
            "Message": message,
        })
        duplicate_rows.append({
            "File": ", ".join(files),
            "index": idx1,
            "index2": idx2,
            "Rows": "",
            "Sample_IDs": "",
        })

    issues_df = pd.DataFrame(all_issues)
    if issues_df.empty:
        issues_df = pd.DataFrame(columns=["File", "Row", "Sample_ID", "Severity", "Category", "Message"])
    amplicon_df = pd.DataFrame(amplicon_rows)
    duplicate_df = pd.DataFrame(duplicate_rows)
    summary_df = pd.DataFrame(file_summaries)

    return {
        "issues_df": issues_df,
        "amplicon_df": amplicon_df,
        "duplicate_df": duplicate_df,
        "summary_df": summary_df,
        "user_ids": sorted(user_ids),
        "files_checked": len(uploaded_files),
        "rows_checked": rows_checked_total,
        "error_count": int((issues_df["Severity"] == "Error").sum()),
        "warning_count": int((issues_df["Severity"] == "Warning").sum()),
        "info_count": int((issues_df["Severity"] == "Info").sum()),
        "cross_dup_combos": cross_dup,
    }


def clean_and_qc_grna_merge(df: pd.DataFrame):
    if "gRNA" not in df.columns:
        return df, {"u_to_t": 0, "invalid_after": 0, "invalid_examples": [], "missing_grna": len(df)}
    grna_orig = df["gRNA"].copy()
    contains_u_mask = grna_orig.astype(str).str.contains(r"[Uu]", regex=True, na=False)
    missing_mask = grna_orig.astype(str).str.strip() == ""
    df["gRNA"] = grna_orig.apply(normalize_grna_for_merge)
    nonempty_mask = df["gRNA"].astype(str).str.len() > 0
    invalid_mask = nonempty_mask & ~df["gRNA"].astype(str).str.match(DNA_ONLY_RE)
    return df, {
        "u_to_t": int(contains_u_mask.sum()),
        "invalid_after": int(invalid_mask.sum()),
        "invalid_examples": df.loc[invalid_mask, "gRNA"].astype(str).unique().tolist()[:5],
        "missing_grna": int(missing_mask.sum()),
    }


def apply_base_editing_rules_merge(df: pd.DataFrame):
    df_out = df.copy()
    if BASE_EDITING_COL not in df_out.columns:
        df_out[BASE_EDITING_COL] = ""
    df_out[BASE_EDITING_COL] = df_out[BASE_EDITING_COL].apply(lambda x: normalize_base_editing_type(clean_cell_merge(x)))

    amp_invalid_examples = []
    hdr_invalid_examples = []
    grna_not_found_examples = []
    base_edit_invalid_examples = []
    base_edit_conflict_examples = []
    reoriented_examples = []

    amp_invalid_count = 0
    hdr_invalid_count = 0
    grna_not_found_count = 0
    base_edit_invalid_count = 0
    base_edit_conflict_count = 0
    reoriented_count = 0

    for idx, row in df_out.iterrows():
        sample_id = clean_cell_merge(row.get("Sample_ID", "")) or f"row {idx + 2}"
        amp_raw = clean_cell_merge(row.get("Amplicon", ""))
        amp_norm = amp_raw.upper().replace("U", "T") if amp_raw else ""
        grna_raw = clean_cell_merge(row.get("gRNA", ""))
        grna_norm = normalize_grna_for_merge(grna_raw) if grna_raw else ""
        hdr_amp_raw = clean_cell_merge(row.get("Expected_HDR_Amplicon", ""))
        hdr_amp_norm = hdr_amp_raw.upper() if hdr_amp_raw else ""
        bet = normalize_base_editing_type(row.get(BASE_EDITING_COL, ""))

        if amp_norm and not DNA_ONLY_RE.fullmatch(amp_norm):
            amp_invalid_count += 1
            if len(amp_invalid_examples) < 5:
                amp_invalid_examples.append(f"{sample_id}: {amp_raw}")

        if hdr_amp_norm and not DNA_ONLY_RE.fullmatch(hdr_amp_norm):
            hdr_invalid_count += 1
            if len(hdr_invalid_examples) < 5:
                hdr_invalid_examples.append(f"{sample_id}: {hdr_amp_raw}")

        if grna_norm:
            grna_rc = reverse_complement(grna_norm)
            grna_in_amp = bool(amp_norm) and (grna_norm in amp_norm)
            grna_rc_in_amp = bool(amp_norm) and (grna_rc in amp_norm)
            if amp_norm and not (grna_in_amp or grna_rc_in_amp):
                grna_not_found_count += 1
                if len(grna_not_found_examples) < 5:
                    grna_not_found_examples.append(sample_id)
            if bet in VALID_BASE_EDITING_TYPES and amp_norm:
                amp_rc = reverse_complement(amp_norm)
                grna_in_amp_rc = grna_norm in amp_rc
                if (not grna_in_amp) and grna_in_amp_rc:
                    df_out.at[idx, "Amplicon"] = amp_rc
                    reoriented_count += 1
                    if len(reoriented_examples) < 5:
                        reoriented_examples.append(sample_id)

        if bet:
            if bet not in VALID_BASE_EDITING_TYPES:
                base_edit_invalid_count += 1
                if len(base_edit_invalid_examples) < 5:
                    base_edit_invalid_examples.append(f"{sample_id}: {bet}")
            else:
                has_conflict = any(clean_cell_merge(row.get(col, "")) != "" for col in BASE_EDITING_CONFLICT_COLS if col in df_out.columns)
                if has_conflict:
                    base_edit_conflict_count += 1
                    if len(base_edit_conflict_examples) < 5:
                        base_edit_conflict_examples.append(sample_id)

    return blankify_merge(df_out), {
        "amp_invalid": amp_invalid_count,
        "amp_invalid_examples": amp_invalid_examples,
        "hdr_invalid": hdr_invalid_count,
        "hdr_invalid_examples": hdr_invalid_examples,
        "grna_not_found": grna_not_found_count,
        "grna_not_found_examples": grna_not_found_examples,
        "base_edit_invalid": base_edit_invalid_count,
        "base_edit_invalid_examples": base_edit_invalid_examples,
        "base_edit_conflict": base_edit_conflict_count,
        "base_edit_conflict_examples": base_edit_conflict_examples,
        "amplicon_reoriented": reoriented_count,
        "amplicon_reoriented_examples": reoriented_examples,
    }


def run_merge(uploaded_files):
    raw_list, exp_list, logs = [], [], []
    file_combos = {}
    grna_qc_logs = []
    base_edit_qc_logs = []
    user_ids = set()

    total_u_to_t = total_invalid = total_missing_grna = 0
    total_invalid_examples = set()
    total_amp_invalid = total_hdr_invalid = total_grna_not_found = total_base_edit_invalid = total_base_edit_conflict = total_amplicon_reoriented = 0
    total_amp_invalid_examples = set()
    total_hdr_invalid_examples = set()
    total_grna_not_found_examples = set()
    total_base_edit_invalid_examples = set()
    total_base_edit_conflict_examples = set()
    total_amplicon_reoriented_examples = set()

    for upload_idx, excel_file in enumerate(uploaded_files):
        file_label = display_file_label(excel_file, upload_idx)
        try:
            df_data = read_data_sheet(excel_file)
        except ValueError:
            logs.append({
                "File": file_label, "Sheet Found": False, "Input Samples": 0,
                "Missing gRNA": 0, "In-file Dup Combos": 0, "Cross-file Dup": False,
                "gRNA U→T": 0, "Non-ATCG (post U→T)": 0,
                "Amplicon Non-ATCG": 0, "gRNA Not Found in Amplicon": 0,
                "Invalid Base_Editing_Type": 0, "Base Editing Conflicts": 0,
                "Amplicons Reoriented": 0,
            })
            continue

        user_ids.update(read_user_ids(excel_file))
        df = normalize_header_columns(blankify_merge(df_data))
        if BASE_EDITING_COL not in df.columns:
            df[BASE_EDITING_COL] = ""

        missing_core_cols = [col for col in CORE_COLS if col not in df.columns]
        for col in missing_core_cols:
            df[col] = ""

        relevant_cols = [c for c in CORE_COLS + QC_OPTIONAL_COLS if c in df.columns]
        mask = pd.Series(False, index=df.index)
        if relevant_cols:
            mask = df[relevant_cols].apply(lambda r: any(clean_cell_merge(v) != "" for v in r), axis=1)
        df_nonblank = blankify_merge(df[mask].copy())

        df_nonblank, qc = clean_and_qc_grna_merge(df_nonblank)
        raw_list.append(df_nonblank)

        dup_details, combos = analyze_duplicates(df_nonblank.assign(_row_num=df_nonblank.index + 2), file_label, [])
        file_combos[file_label] = combos
        in_dup_count = len(dup_details)

        df_processed, qc_be = apply_base_editing_rules_merge(df_nonblank)
        exp_list.append(df_processed)

        total_u_to_t += qc["u_to_t"]
        total_invalid += qc["invalid_after"]
        total_missing_grna += qc["missing_grna"]
        total_invalid_examples.update(qc["invalid_examples"])
        total_amp_invalid += qc_be["amp_invalid"]
        total_amp_invalid_examples.update(qc_be["amp_invalid_examples"])
        total_hdr_invalid += qc_be["hdr_invalid"]
        total_hdr_invalid_examples.update(qc_be["hdr_invalid_examples"])
        total_grna_not_found += qc_be["grna_not_found"]
        total_grna_not_found_examples.update(qc_be["grna_not_found_examples"])
        total_base_edit_invalid += qc_be["base_edit_invalid"]
        total_base_edit_invalid_examples.update(qc_be["base_edit_invalid_examples"])
        total_base_edit_conflict += qc_be["base_edit_conflict"]
        total_base_edit_conflict_examples.update(qc_be["base_edit_conflict_examples"])
        total_amplicon_reoriented += qc_be["amplicon_reoriented"]
        total_amplicon_reoriented_examples.update(qc_be["amplicon_reoriented_examples"])

        grna_qc_logs.append({
            "File": file_label,
            "Missing gRNA": qc["missing_grna"],
            "gRNA U→T": qc["u_to_t"],
            "Non-ATCG (post U→T)": qc["invalid_after"],
            "Examples (up to 5)": ", ".join(qc["invalid_examples"]) if qc["invalid_examples"] else "",
        })
        base_edit_qc_logs.append({
            "File": file_label,
            "Amplicon Non-ATCG": qc_be["amp_invalid"],
            "Expected_HDR_Amplicon Non-ATCG": qc_be["hdr_invalid"],
            "gRNA Not Found in Amplicon": qc_be["grna_not_found"],
            "Invalid Base_Editing_Type": qc_be["base_edit_invalid"],
            "Base Editing Conflicts": qc_be["base_edit_conflict"],
            "Amplicons Reoriented": qc_be["amplicon_reoriented"],
        })
        logs.append({
            "File": file_label, "Sheet Found": True, "Input Samples": len(df_nonblank),
            "Missing gRNA": qc["missing_grna"], "In-file Dup Combos": in_dup_count, "Cross-file Dup": False,
            "gRNA U→T": qc["u_to_t"], "Non-ATCG (post U→T)": qc["invalid_after"],
            "Amplicon Non-ATCG": qc_be["amp_invalid"],
            "Expected_HDR_Amplicon Non-ATCG": qc_be["hdr_invalid"],
            "gRNA Not Found in Amplicon": qc_be["grna_not_found"],
            "Invalid Base_Editing_Type": qc_be["base_edit_invalid"],
            "Base Editing Conflicts": qc_be["base_edit_conflict"],
            "Amplicons Reoriented": qc_be["amplicon_reoriented"],
        })

    cross_dup = cross_file_duplicates(file_combos)
    for entry in logs:
        entry["Cross-file Dup"] = any(entry["File"] in files for files in cross_dup.values())

    raw_df = blankify_merge(pd.concat(raw_list, ignore_index=True)) if raw_list else pd.DataFrame(columns=FINAL_COLS)
    expanded_df = blankify_merge(pd.concat(exp_list, ignore_index=True)) if exp_list else pd.DataFrame(columns=FINAL_COLS)

    log_df = pd.DataFrame(logs)
    if not log_df.empty:
        totals = {"File": "Total"}
        for col in log_df.columns:
            if col not in ("File", "Sheet Found", "Cross-file Dup"):
                totals[col] = int(pd.to_numeric(log_df[col], errors="coerce").fillna(0).sum())
        log_rows = pd.concat([log_df, pd.DataFrame([totals])], ignore_index=True)
    else:
        log_rows = pd.DataFrame(columns=["File"])

    return {
        "raw_df": raw_df,
        "expanded_df": expanded_df,
        "log_rows": log_rows,
        "grna_qc_logs": pd.DataFrame(grna_qc_logs) if grna_qc_logs else pd.DataFrame(),
        "base_edit_qc_logs": pd.DataFrame(base_edit_qc_logs) if base_edit_qc_logs else pd.DataFrame(),
        "grna_qc_totals": {
            "missing_grna": total_missing_grna,
            "u_to_t": total_u_to_t,
            "invalid_after": total_invalid,
            "invalid_examples": sorted(list(total_invalid_examples))[:5],
        },
        "base_edit_qc_totals": {
            "amp_invalid": total_amp_invalid,
            "amp_invalid_examples": sorted(list(total_amp_invalid_examples))[:5],
            "hdr_invalid": total_hdr_invalid,
            "hdr_invalid_examples": sorted(list(total_hdr_invalid_examples))[:5],
            "grna_not_found": total_grna_not_found,
            "grna_not_found_examples": sorted(list(total_grna_not_found_examples))[:5],
            "base_edit_invalid": total_base_edit_invalid,
            "base_edit_invalid_examples": sorted(list(total_base_edit_invalid_examples))[:5],
            "base_edit_conflict": total_base_edit_conflict,
            "base_edit_conflict_examples": sorted(list(total_base_edit_conflict_examples))[:5],
            "amplicon_reoriented": total_amplicon_reoriented,
            "amplicon_reoriented_examples": sorted(list(total_amplicon_reoriented_examples))[:5],
        },
        "cross_dup_combos": cross_dup,
        "user_ids": sorted(user_ids),
    }


def render_qc_results(results):
    st.subheader("QC Summary")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Files Checked", results["files_checked"])
    c2.metric("Rows Checked", results["rows_checked"])
    c3.metric("Errors", results["error_count"], help="Errors must be corrected before the sheet is ready.")
    c4.metric("Warnings", results["warning_count"], help="Warnings should be reviewed. They may or may not need correction.")
    c5.metric("Info", results["info_count"], help="Info items are for awareness only and do not require correction.")

    if results["user_ids"]:
        st.info(f"Unique User IDs found: {', '.join(results['user_ids'])}")

    if not results["summary_df"].empty:
        st.subheader("Per-file Status")
        st.dataframe(results["summary_df"], use_container_width=True, hide_index=True)

    if not results["issues_df"].empty:
        st.subheader("QC Issues")
        st.dataframe(results["issues_df"], use_container_width=True, hide_index=True)
    else:
        st.success("No QC issues found.")

    if not results["amplicon_df"].empty:
        st.subheader("Unique Amplicon Length Summary")
        display_cols = ["File", "Amplicon Label", "Length", "Occurrence Count", "Sequence"]
        st.dataframe(results["amplicon_df"][display_cols], use_container_width=True, hide_index=True)

    if not results["duplicate_df"].empty:
        st.subheader("Duplicate Index/index2 Details")
        st.dataframe(results["duplicate_df"], use_container_width=True, hide_index=True)


def render_merge_results(results, prefix):
    final_df = ensure_columns_and_order(results["expanded_df"], FINAL_COLS)

    grna_totals = results["grna_qc_totals"]
    if grna_totals["missing_grna"] > 0:
        st.warning(f"⚠️Found **{grna_totals['missing_grna']}** row(s) with missing gRNA. gRNA is required for CRISPResso.")
    if grna_totals["u_to_t"] > 0:
        st.info(f"gRNA cleanup: Converted **{grna_totals['u_to_t']}** entries from U→T (case-insensitive).")
    if grna_totals["invalid_after"] > 0:
        examples = ", ".join(grna_totals["invalid_examples"])
        msg = f"Found **{grna_totals['invalid_after']}** gRNA entry(ies) with non-A/T/C/G characters after U→T conversion."
        if examples:
            msg += f" Examples: {examples}"
        st.warning(msg)

    be_totals = results["base_edit_qc_totals"]
    if be_totals["amp_invalid"] > 0:
        examples = ", ".join(be_totals["amp_invalid_examples"])
        msg = f"⚠️Found **{be_totals['amp_invalid']}** Amplicon entry(ies) with non-A/T/C/G characters."
        if examples:
            msg += f" Examples: {examples}"
        st.warning(msg)
    if be_totals["hdr_invalid"] > 0:
        examples = ", ".join(be_totals["hdr_invalid_examples"])
        msg = f"⚠️Found **{be_totals['hdr_invalid']}** Expected_HDR_Amplicon entry(ies) with non-A/T/C/G characters."
        if examples:
            msg += f" Examples: {examples}"
        st.warning(msg)
    if be_totals["grna_not_found"] > 0:
        examples = ", ".join(be_totals["grna_not_found_examples"])
        msg = f"⚠️Found **{be_totals['grna_not_found']}** row(s) where gRNA was not found in Amplicon and reverse complement(gRNA) was also not found."
        if examples:
            msg += f" Examples: {examples}"
        st.warning(msg)
    if be_totals["base_edit_invalid"] > 0:
        examples = ", ".join(be_totals["base_edit_invalid_examples"])
        msg = f"⚠️Found **{be_totals['base_edit_invalid']}** row(s) with invalid Base_Editing_Type. Allowed values are ABE, CBE, or BOTH."
        if examples:
            msg += f" Examples: {examples}"
        st.warning(msg)
    if be_totals["base_edit_conflict"] > 0:
        examples = ", ".join(be_totals["base_edit_conflict_examples"])
        msg = f"⚠️Found **{be_totals['base_edit_conflict']}** base-editing row(s) that also contain HDR/window/ngRNA fields."
        if examples:
            msg += f" Examples: {examples}"
        st.warning(msg)
    if be_totals["amplicon_reoriented"] > 0:
        examples = ", ".join(be_totals["amplicon_reoriented_examples"])
        msg = f"Re-oriented **{be_totals['amplicon_reoriented']}** Amplicon row(s) so gRNA is in the same sense for base-editing samples."
        if examples:
            msg += f" Examples: {examples}"
        st.info(msg)

    if not results["log_rows"].empty and "In-file Dup Combos" in results["log_rows"].columns:
        per_file_dup_df = results["log_rows"][results["log_rows"]["File"] != "Total"].copy()
        dup_mask = pd.to_numeric(per_file_dup_df["In-file Dup Combos"], errors="coerce").fillna(0) > 0
        if dup_mask.any():
            dup_rows = per_file_dup_df.loc[dup_mask, ["File", "In-file Dup Combos"]]
            details = ", ".join(
                f"{row['File']} ({int(pd.to_numeric(row['In-file Dup Combos'], errors='coerce'))})"
                for _, row in dup_rows.head(5).iterrows()
            )
            total_dup_files = int(dup_mask.sum())
            st.warning(f"⚠️Found in-file duplicate index/index2 combos in **{total_dup_files}** file(s). Examples: {details}")

    if results["cross_dup_combos"]:
        combos_str = ", ".join(f"{i}/{j}" for i, j in results["cross_dup_combos"])
        st.warning(f"⚠️Cross-file duplicate index combos: {combos_str}")

    hdr_mask = final_df["Expected_HDR_Amplicon"].astype(str).str.strip().ne("")
    qwc_blank = final_df["Quantification_Window_Coordinates"].astype(str).str.strip().eq("")
    prob_mask = hdr_mask & qwc_blank
    if prob_mask.any():
        n_prob = int(prob_mask.sum())
        examples = final_df.loc[prob_mask, "Sample_ID"].astype(str).head(5).tolist()
        st.warning(f"⚠️{n_prob} row(s) have Expected_HDR_Amplicon but missing Quantification_Window_Coordinates. Examples: {', '.join(examples)}")

    if results["user_ids"]:
        st.info(f"Merged Unique User IDs: {', '.join(results['user_ids'])}")

    st.subheader("Merge Log")
    st.table(results["log_rows"])

    if not results["grna_qc_logs"].empty:
        with st.expander("Show per-file gRNA QC details"):
            st.table(results["grna_qc_logs"])
    if not results["base_edit_qc_logs"].empty:
        with st.expander("Show per-file Base Editing / Amplicon QC details"):
            st.table(results["base_edit_qc_logs"])

    excel_buf = BytesIO()
    write_excel_with_blanks(final_df, excel_buf, SHEET_NAME, results["user_ids"])
    st.download_button(
        "📥 Download merged Excel (CRISPResso)",
        data=excel_buf,
        file_name=out_excel,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    csv_buf = StringIO()
    csv_buf.write("[Header]\n")
    csv_buf.write(f"Experiment Name,{prefix}\n")
    csv_buf.write(f"Date,{today.month}/{today.day}/{today.year}\n")
    csv_buf.write("Workflow,GenerateFASTQ\n")
    csv_buf.write("[Reads]\n300\n")
    csv_buf.write("[Settings]\n")
    csv_buf.write("[Data]\n")
    csv_buf.write("Sample_ID,Sample_Name,I7_Index_ID,index,I5_INDEX_ID,index2,Sample_Project,Description\n")
    for _, row in results["raw_df"].iterrows():
        csv_buf.write(
            f"{row['Sample_ID']},,{row['I7_Index_ID']},{row['index']},{row['I5_Index_ID']},{row['index2']},,\n"
        )
    st.download_button(
        "📥 Download Miseq CSV",
        data=csv_buf.getvalue().encode("utf-8"),
        file_name=out_csv,
        mime="text/csv",
    )

    st.subheader("Merged Data Preview")
    st.dataframe(final_df, use_container_width=True)


if run_clicked:
    if not uploaded_files:
        st.warning("⚠️Please upload at least one file before running.")
    elif mode == "QC Screening":
        state.qc_results = run_qc(uploaded_files)
        state.merge_results = None
    else:
        state.merge_results = run_merge(uploaded_files)
        state.qc_results = None

if mode == "QC Screening" and state.qc_results is not None:
    render_qc_results(state.qc_results)
elif mode == "File Merge" and state.merge_results is not None:
    render_merge_results(state.merge_results, prefix)
